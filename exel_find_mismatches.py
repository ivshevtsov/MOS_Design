from typing import Set, Any

from openpyxl import load_workbook

#directory
HOME = 'C:/Users/SH/PycharmProjects/Vesta_IoT/DATA/EXEL_convert'

#file name
File = 'CHIP_PADS_V1.xlsx'

#load document
wb = load_workbook(f'{HOME}/{File}')

#number of sheet
N_sheet = 1

#number rows and colums
n_rows = 61
n_columns = 61

#separator
sep = '	'

##------------------------------------------------##
##------------------------------------------------##

#get sheet names
sheetnames = wb.sheetnames
#get sheet data
sheet = wb[sheetnames[N_sheet-1]]

#list elements
List_elements = []

#work with sets

chip_pads: set[Any] = set()
chip_pads_table=set()

for row in range(2, n_rows + 2):
    for column in range(2, n_columns +2):
        Value = sheet.cell(row=row, column=column).value
        chip_pads.add(Value)

print(chip_pads)
print(len(chip_pads))
print('APP_INT_0' in chip_pads)
print('GNSS_GPMC_A_2' in chip_pads)
print('MODEM_GND_ADC' in chip_pads)


#read correctly string
string='GPS_DAT_, GLO_DAT_'

def read_bus(word):
    bus_set=set()
    if word[-1]!= ']' :
        bus_set.add(word)
    else:
        low=int(word[-2])
        high = int(word[-4])

        for i in range(low, high+1):
            string=word[:-5] + '[' + str(i) + ']'
            bus_set.add(string)
    return bus_set


def work_with_one_cell(string):
    set_cell = set()
    split_string = string.split(',') #separate string
    n_string = len(split_string) #number of string
    if n_string>1:
        for i in split_string:
            sum=read_bus(i)
            set_cell.update(sum)
    else:
        set_cell.add(string)

    print(set_cell)


work_with_one_cell(string)








