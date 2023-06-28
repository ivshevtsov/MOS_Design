from openpyxl import load_workbook

##------------------------------------------------##
##------------------------------------------------##
def get_key(d, value):
    for k, v in d.items():
        if v == int(value):
            return k

#directory
HOME = 'Files/Exel_reverse'

#file name
File = 'VestaPinout_400_0908'

#load document
wb = load_workbook(f'{HOME}/{File}.xlsx')

#number of sheet
N_sheet = 5

#number rows and colums
n_rows = 20
n_columns = 20

#separator
sep = '	'

##------------------------------------------------##
##------------------------------------------------##

#get sheet names
sheetnames = wb.sheetnames
#get sheet data
sheet = wb[sheetnames[N_sheet-1]]

#list elements
List_elements = []

#work with dictionares

dict_sample = {}

A=65
k=0
alphabets_in_capital=[]
for i in range(A,A+n_rows):
    k+=1
    dict_sample[chr(i)] = k
print(dict_sample)


with open(f'{HOME}/{File}.txt', "w") as file:
    for row in range(1, n_rows+1):
        for column in range(1, n_columns+1):
            #row_x = sheet.cell(row=row, column=1).value
            #column_x = sheet.cell(row=1, column=column).value
            Value = sheet.cell(row=row, column=column).value
            coordinate = sheet.cell(row=row, column=column).coordinate
            first=coordinate[0]
            second=coordinate[1:]

            #String = f'{row_x}'+f'{column_x}'+ f'{sep}'  f'{Value}'
            String = f'{get_key(dict_sample, second)}'+ f'{dict_sample[first]}'+f'{sep}'  f'{Value}'


            #create list elements for compare
            List_elements.append(String)
            #print(String)
            file.write(f'{String} \n')
    file.close()


#Compare Pins
'''
with open(f'{HOME}/Error.txt', 'w') as file_1:
    with open(f'{HOME}/Altium.txt') as file:
        for line in file:
            if line.strip() not in List_elements:
                print(line.strip())
                file_1.write(f'{line.strip()}\n')
    file.close()
file_1.close()

'''









