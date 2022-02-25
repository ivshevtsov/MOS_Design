from openpyxl import load_workbook

#directory
HOME = 'Files/Exel_reverse'

#file name
File = 'SILICAT_SUB_BALLS_fin'

#load document
wb = load_workbook(f'{HOME}/{File}.xlsx')

#number of sheet
N_sheet = 1

#number rows and colums
n_rows = 34
n_columns = 34

#separator
sep = '	'

##------------------------------------------------##
##------------------------------------------------##

#get sheet names
sheetnames = wb.sheetnames
#get sheet data
sheet = wb[sheetnames[N_sheet-1]]

#list elements
List_elements = []


with open(f'{HOME}/{File}.txt', "w") as file:
    for row in range(2, n_rows+1):
        for column in range(2, n_columns+1):
            row_x = sheet.cell(row=row, column=1).value
            column_x = sheet.cell(row=1, column=column).value
            Value = sheet.cell(row=row, column=column).value
            String = f'{row_x}'+f'{column_x}'+ f'{sep}'  f'{Value}'

            #create list elements for compare
            List_elements.append(String)
            #print(String)
            file.write(f'{String} \n')
    file.close()


#Compare Pins

with open(f'{HOME}/Error.txt', 'w') as file_1:
    with open(f'{HOME}/Altium.txt') as file:
        for line in file:
            if line.strip() not in List_elements:
                print(line.strip())
                file_1.write(f'{line.strip()}\n')
    file.close()
file_1.close()








