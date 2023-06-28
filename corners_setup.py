import numpy as np
from openpyxl import load_workbook




Model_File = 'Modelfile::/home/DATA/PDK/cadence_oa/t-n40-cm-sp-001-k3_2_0_2a_20140217_new/models/spectre/crn40lp_2d5_v2d0_2_shrink0d9_embedded_usage.scs'
Test = 't Test::Vesta_Filters_40:Test_RX_LPF_V1_woTR:1'

Home = 'C:/Users/SH/PycharmProjects/Vesta_IoT/DATA/Filters_corners'
File = 'corners_file.csv'

CSV_CORNER = np.genfromtxt(f'{Home}/{File}', delimiter=',', dtype='str')


#load document
wb = load_workbook(f'{Home}/{File}.xlsx')
#number of sheet
N_sheet = 1

##------------------------------------------------##
##------------------------------------------------##

#get sheet names
sheetnames = wb.sheetnames
#get sheet data
sheet = wb[sheetnames[N_sheet-1]]




